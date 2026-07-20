import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
import logging

logger = logging.getLogger(__name__)

def style_excel_sheet(file_path: str):
    """
    Applies professional styling to an Excel workbook:
    - Bold header row with a pleasant blue fill and white text.
    - Centered content where appropriate.
    - Grid lines enabled.
    - Auto-fitted column widths based on content.
    """
    if not os.path.exists(file_path):
        return
        
    try:
        wb = openpyxl.load_workbook(file_path)
        
        # Color palettes
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=10)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.views.sheetView[0].showGridLines = True
            
            # Style header row
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 28
            
            # Style data rows and auto-fit columns
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 20
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = regular_font
                    
                    # Align scores and years to center
                    col_header = ws.cell(row=1, column=col).value
                    if col_header and any(h in str(col_header).lower() for h in ["score", "year", "pages", "count"]):
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            # Auto-fit columns
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val = cell.value
                    if val is not None:
                        max_len = max(max_len, len(str(val)))
                # Set reasonable constraints
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)
                
        wb.save(file_path)
        wb.close()
    except Exception as e:
        logger.error(f"Error styling Excel report {file_path}: {e}")

class ExcelReporter:
    """
    Exports matched, unmatched, and duplicate sheets to beautifully formatted Excel sheets.
    """
    @staticmethod
    def export_matching_results(results: dict, output_dir: str = ".") -> tuple:
        """
        Saves 'matched' and 'unmatched' results to Excel sheets.
        """
        os.makedirs(output_dir, exist_ok=True)
        
        matched_path = os.path.join(output_dir, "report.xlsx")
        unmatched_path = os.path.join(output_dir, "unmatched.xlsx")
        
        # 1. Matched report
        matched_data = results.get("matched", [])
        if matched_data:
            df_matched = pd.DataFrame(matched_data)
            # Rename columns to be user friendly
            df_matched.columns = [
                "Target Original Citation", "Target Inferred Title", "Target Inferred DOI",
                "Matched File Name", "Matched File Path", "Matched Document Title",
                "Matched Authors", "Matched DOI", "Matched Journal", "Matched Year",
                "Document Type", "Match Score (%)", "Matching Method"
            ]
        else:
            df_matched = pd.DataFrame(columns=["Target Original Citation", "Target Inferred Title", "Status"])
            df_matched.loc[0] = ["No matches found", "", "Unmatched"]
            
        df_matched.to_excel(matched_path, index=False)
        style_excel_sheet(matched_path)
        
        # 2. Unmatched report
        unmatched_data = results.get("unmatched", [])
        if unmatched_data:
            df_unmatched = pd.DataFrame(unmatched_data)
            df_unmatched.columns = [
                "Target Original Citation", "Target Inferred Title", "Target Inferred DOI",
                "Best Index Candidate Title", "Best Candidate Similarity Score (%)"
            ]
        else:
            df_unmatched = pd.DataFrame(columns=["Status"])
            df_unmatched.loc[0] = ["All targets matched successfully!"]
            
        df_unmatched.to_excel(unmatched_path, index=False)
        style_excel_sheet(unmatched_path)
        
        return matched_path, unmatched_path

    @staticmethod
    def export_duplicates_report(exact_dups: list, potential_dups: list, output_dir: str = ".") -> str:
        """
        Saves exact and potential duplicates to a single workbook with separate sheets.
        """
        os.makedirs(output_dir, exist_ok=True)
        file_path = os.path.join(output_dir, "duplicates.xlsx")
        
        # Prepare exact duplicates data
        exact_rows = []
        for i, group in enumerate(exact_dups, 1):
            for doc in group:
                exact_rows.append({
                    "Duplicate Group": f"Group {i}",
                    "File Name": doc["file_name"],
                    "Title": doc["title"],
                    "Authors": doc["authors"],
                    "DOI": doc["doi"],
                    "Journal": doc["journal"],
                    "Year": doc["year"],
                    "File Path": doc["file_path"]
                })
                
        # Prepare potential duplicates data
        potential_rows = []
        for i, group in enumerate(potential_dups, 1):
            for doc in group:
                potential_rows.append({
                    "Duplicate Group": f"Group {i}",
                    "File Name": doc["file_name"],
                    "Title": doc["title"],
                    "Authors": doc["authors"],
                    "DOI": doc["doi"],
                    "Journal": doc["journal"],
                    "Year": doc["year"],
                    "File Path": doc["file_path"]
                })
                
        df_exact = pd.DataFrame(exact_rows) if exact_rows else pd.DataFrame(columns=["Message"])
        if not exact_rows:
            df_exact.loc[0] = ["No exact duplicates found by file hash."]
            
        df_potential = pd.DataFrame(potential_rows) if potential_rows else pd.DataFrame(columns=["Message"])
        if not potential_rows:
            df_potential.loc[0] = ["No potential duplicates found by identical titles."]
            
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df_exact.to_excel(writer, sheet_name="Exact (Hash) Duplicates", index=False)
            df_potential.to_excel(writer, sheet_name="Potential (Title) Duplicates", index=False)
            
        style_excel_sheet(file_path)
        return file_path
